import openpyxl
import os
import json
imagedir = 'S:\\skill.xlsx'
skillDir = 'S:\\workspace\\idris\\code\\combat_engine\\ts\\src\\skill'

buffExcal = 'S:\\buff.xlsx'
buffClassDir = 'S:\\workspace\\idris\\code\\combat_engine\\ts\\src\\buff'

tagExcal = 'S:\\tag.xlsx'
tagClassDir = 'S:\\workspace\\idris\\code\\combat_engine\\ts\\src\\tag'

def readSkillExcel(excelFile, skillDir):
    workbook = openpyxl.load_workbook(excelFile)
    worksheet = workbook['skill']
    col3_values = [item.value for item in list(worksheet.columns)[8]]
    col3_values = list(set(col3_values))
    print(len(col3_values))
    explore_list = get_filelist(skillDir, [])
    for fileName in explore_list:
        baseName = os.path.basename(fileName)
        sourceSp = baseName.split('.', -1)[0]
        if sourceSp not in col3_values:
            if sourceSp != "Skills" and sourceSp != "SkillShoot" and sourceSp != "SkillSample":
                print("remove "+sourceSp)

def readBuffExcel(excelFile, skillDir):
    workbook = openpyxl.load_workbook(excelFile)
    worksheet = workbook['buff']
    col3_values = [item.value for item in list(worksheet.columns)[3]]
    col3_values = list(set(col3_values))
    print(len(col3_values))
    explore_list = get_filelist(skillDir, [])
    for fileName in explore_list:
        baseName = os.path.basename(fileName)
        sourceSp = baseName.split('.', -1)[0]
        if sourceSp not in col3_values:
            if sourceSp != "Buffs" and sourceSp != "BuffSample" and sourceSp != "IBuffTaunted":
                print("remove "+sourceSp)
                # os.remove(fileName)

def readTagExcel(excelFile, skillDir):
    workbook = openpyxl.load_workbook(excelFile)
    worksheet = workbook['Sheet1']
    col3_values = [item.value for item in list(worksheet.columns)[3]]
    col3_values = list(set(col3_values))
    print(len(col3_values))
    explore_list = get_filelist(skillDir, [])
    # for fileName in explore_list:
    #     baseName = os.path.basename(fileName)
    #     sourceSp = baseName.split('.', -1)[0]
    #     if sourceSp not in col3_values:
    #         if sourceSp != "Tags":
    #             print("remove "+sourceSp)
    #             # os.remove(fileName)
    for value in col3_values:
        newfile = tagClassDir+"\\"+value+".ts"
        if(not os.path.lexists(newfile)):
            # open(newfile, "w")
            print(value + " not exist")


def get_filelist(dir, Filelist):
    newDir = dir
    if os.path.isfile(dir):
        Filelist.append(dir)
    elif os.path.isdir(dir):
        for s in os.listdir(dir):
            # 如果需要忽略某些文件夹，使用以下代码
            # if s == "Skill" or s == "SkillShoot" or s == "SkillSample":
            #     continue

            newDir = os.path.join(dir, s)

            get_filelist(newDir, Filelist)
    return Filelist


if __name__ == '__main__':
    print("packer start")
    # readSkillExcel(imagedir, skillDir)
    # readBuffExcel(buffExcal, buffClassDir)
    readTagExcel(tagExcal, tagClassDir)
    print("packer end")